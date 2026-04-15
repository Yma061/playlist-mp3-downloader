import yt_dlp
import openpyxl
import os
import re
import sys
import time
import random
import socket

def _ffmpeg_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return None

EXCEL_PATH = "playlist excel/10 - 1er mai 2026 prévisions_v12 du 04_01_2026.xlsx"
SHEET_NAME = "Deezer"
TITRE_COLUMN = 5
FIRST_ROW = 7
OUTPUT_FOLDER = "playlists"
PLAYLIST_NAME = "1er mai 2026"
MAX_RETRIES = 3


def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "", filename)


def is_connected():
    try:
        socket.setdefaulttimeout(3)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect(("8.8.8.8", 53))
        return True
    except Exception:
        return False


def wait_for_connection(print_fn=print):
    if is_connected():
        return
    print_fn("  Connexion perdue — attente du retour internet...")
    while not is_connected():
        time.sleep(5)
    print_fn("  Connexion rétablie, reprise du téléchargement.")


def get_sheets(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    return wb.sheetnames


def get_tracks_from_excel(excel_path, sheet_name, titre_col, first_row):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    tracks = []
    for row in ws.iter_rows(min_row=first_row, values_only=True):
        titre = row[titre_col - 1]
        if titre and isinstance(titre, str) and titre.strip():
            tracks.append(titre.strip())
    return tracks


def download_tracks(tracks, output_folder, playlist_name, on_progress=None):
    dest = os.path.join(output_folder, sanitize_filename(playlist_name))
    os.makedirs(dest, exist_ok=True)

    total = len(tracks)
    failed = []
    print(f"{total} titres à télécharger dans '{dest}'\n")

    for index, titre in enumerate(tracks, start=1):
        num = str(index).zfill(len(str(total)))
        out_path = os.path.join(dest, f"{num} - %(title)s.%(ext)s")

        ydl_opts = {
            'format': 'bestaudio/best',
            'postprocessors': [{
                'key': 'FFmpegExtractAudio',
                'preferredcodec': 'mp3',
                'preferredquality': '192',
            }],
            'outtmpl': {'default': out_path},
            'nooverwrites': True,
            'quiet': True,
            'no_warnings': True,
            'default_search': 'ytsearch1',
            'socket_timeout': 20,
            **({'ffmpeg_location': _ffmpeg_path()} if _ffmpeg_path() else {}),
        }

        print(f"[{index}/{total}] {titre}")
        if on_progress:
            on_progress(index - 1, total, titre)

        success = False
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    ydl.download([titre])
                success = True
                break
            except Exception as e:
                err = str(e).lower()
                is_network = any(k in err for k in ("timed out", "connection", "network", "read timed"))
                if is_network:
                    wait_for_connection()
                if attempt < MAX_RETRIES:
                    print(f"  Tentative {attempt}/{MAX_RETRIES} échouée, nouvel essai...")
                    time.sleep(3)
                else:
                    print(f"  ERREUR après {MAX_RETRIES} tentatives : {e}")

        if not success:
            failed.append(f"{num} - {titre}")

        if on_progress:
            on_progress(index, total, titre)

        if index < total:
            pause = random.uniform(3, 8)
            print(f"  Pause {pause:.1f}s...")
            time.sleep(pause)

    print("\nTéléchargement terminé !")

    if failed:
        log_path = os.path.join(dest, "non_trouves.txt")
        with open(log_path, "w", encoding="utf-8") as f:
            f.write(f"{len(failed)} titre(s) non trouvé(s) :\n\n")
            f.write("\n".join(failed))
        print(f"\n{len(failed)} titre(s) non trouvé(s) — voir '{log_path}'")
    else:
        print("Tous les titres ont été téléchargés avec succès !")

    return dest, failed


if __name__ == "__main__":
    tracks = get_tracks_from_excel(EXCEL_PATH, SHEET_NAME, TITRE_COLUMN, FIRST_ROW)
    download_tracks(tracks, OUTPUT_FOLDER, PLAYLIST_NAME)
